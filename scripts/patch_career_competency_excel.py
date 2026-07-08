"""
Xóa các dòng mapping career–competency sai trong Excel nguồn.

Chạy: python scripts/patch_career_competency_excel.py
Sau đó: python scripts/ingest.py --reset-graph
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_ROOT / "data" / "bộ dữ liệu.xlsx"

# map_id hoặc (career_code, item_code)
REMOVE_MAP_IDS: set[str] = {"M00248"}
REMOVE_PAIRS: set[tuple[str, str]] = {("DA", "L_DART")}


def patch_sheet(xlsx_path: Path) -> int:
    wb = load_workbook(xlsx_path)
    if "career_competency_map" not in wb.sheetnames:
        raise SystemExit("Sheet career_competency_map không tồn tại")

    ws = wb["career_competency_map"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        idx_map = headers.index("map_id")
        idx_career = headers.index("career_code")
        idx_item = headers.index("item_code")
    except ValueError as exc:
        raise SystemExit(f"Thiếu cột bắt buộc: {exc}") from exc

    removed = 0
    for row_idx in range(len(rows) - 1, 0, -1):
        row = rows[row_idx]
        map_id = str(row[idx_map] or "").strip()
        career = str(row[idx_career] or "").strip()
        item = str(row[idx_item] or "").strip()
        if map_id in REMOVE_MAP_IDS or (career, item) in REMOVE_PAIRS:
            ws.delete_rows(row_idx + 1)
            removed += 1
            print(f"Removed row {row_idx + 1}: map_id={map_id} career={career} item={item}")

    if removed:
        wb.save(xlsx_path)
        print(f"Saved ({removed} row(s) removed)")
    else:
        print("No rows removed (already patched?)")
    wb.close()
    return removed


if __name__ == "__main__":
    patch_sheet(DEFAULT_XLSX)
