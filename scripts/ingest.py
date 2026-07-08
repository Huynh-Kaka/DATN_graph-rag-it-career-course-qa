"""
Nạp dữ liệu từ Excel vào Neo4j: mô hình nút theo loại (Knowledge, Tool, …),
quan hệ NEED_* / TEACH_* từ type_code, và thuộc tính `color` (hex) cho từng nhãn.

Chạy: python scripts/ingest.py [--xlsx-path ...] [--batch-size 500] [--reset-graph]
"""

from __future__ import annotations

import argparse
import json
import os
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from dotenv import load_dotenv
from neo4j import GraphDatabase
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX_PATH = PROJECT_ROOT / "data" / "bộ dữ liệu.xlsx"
RELATION_SUPPLEMENT_PATH = PROJECT_ROOT / "data" / "competency_relation_supplement.jsonl"

# --- Bảng màu (hex) theo nhãn — đồng bộ với design/neo4j_browser_palette.grass ---
NODE_COLORS: Dict[str, str] = {
    "Taxonomy": "#F1C40F",
    "Career": "#E74C3C",
    "Subject": "#95A5A6",
    "Program": "#BB8FCE",
    "Subtitle": "#AEB6BF",
    "Platform": "#D4AC6E",
    "Knowledge": "#E67E22",
    "Industry": "#1F4E79",
    "Course": "#58D68D",
    "Level": "#1E8449",
    "Instructor": "#E91E63",
    "Tool": "#1ABC9C",
    "Framework": "#9B59B6",
    "ProgrammingLanguage": "#00BCD4",
    "Softskill": "#2ECC71",
    "Certification": "#8E44AD",
    "Website": "#BDC3C7",
    "Organization": "#7F8C8D",
    "CompetencyType": "#CCD1D1",
}

TYPE_CODE_TO_LABEL: Dict[str, str] = {
    "CT_LANG": "ProgrammingLanguage",
    "CT_FRAM": "Framework",
    "CT_PLAT": "Platform",
    "CT_TOOL": "Tool",
    "CT_KNOW": "Knowledge",
    "CT_SOFT": "Softskill",
    "CT_CERT": "Certification",
}

CAREER_REL_BY_TYPE: Dict[str, str] = {
    "CT_LANG": "NEED_LANG",
    "CT_FRAM": "NEED_FRAM",
    "CT_PLAT": "NEED_PLAT",
    "CT_TOOL": "NEED_TOOL",
    "CT_KNOW": "NEED_KNOW",
    "CT_SOFT": "NEED_SOFT",
    "CT_CERT": "NEED_CERT",
}

# (career_code, item_code) — loại bỏ mapping sai nguồn Excel trước khi MERGE
CAREER_COMPETENCY_DENYLIST: set[tuple[str, str]] = {
    ("DA", "L_DART"),  # Data Analyst không cần Dart
}

TEACH_REL_BY_TYPE: Dict[str, str] = {
    "CT_LANG": "TEACH_LANG",
    "CT_FRAM": "TEACH_FRAM",
    "CT_PLAT": "TEACH_PLAT",
    "CT_TOOL": "TEACH_TOOL",
    "CT_KNOW": "TEACH_KNOW",
    "CT_SOFT": "TEACH_SOFT",
    "CT_CERT": "TEACH_CERT",
}

# Sheet Excel -> nhãn Neo4j (phải khớp TYPE_CODE_TO_LABEL qua type_code trong từng dòng)
COMPETENCY_SHEETS: List[Tuple[str, str]] = [
    ("programming_language", "ProgrammingLanguage"),
    ("framework", "Framework"),
    ("platform", "Platform"),
    ("tool", "Tool"),
    ("knowledge", "Knowledge"),
    ("softskill", "Softskill"),
    ("certification", "Certification"),
]


def read_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[Dict[str, object]]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return []

    raw_headers = rows[0]
    headers: List[str] = []
    header_positions: List[int] = []
    for idx, header in enumerate(raw_headers):
        if header is None:
            continue
        header_text = str(header).strip()
        if not header_text:
            continue
        headers.append(header_text)
        header_positions.append(idx)

    parsed_rows: List[Dict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        item = {}
        for out_idx, col_idx in enumerate(header_positions):
            value = row[col_idx] if col_idx < len(row) else None
            item[headers[out_idx]] = value
        if any(v is not None and str(v).strip() != "" for v in item.values()):
            parsed_rows.append(item)

    workbook.close()
    return parsed_rows


def load_relation_supplement(path: Path | None = None) -> List[Dict[str, object]]:
    """Optional JSONL rows merged into competency_relation (Phase 0 extension)."""
    p = path or RELATION_SUPPLEMENT_PATH
    if not p.is_file():
        return []
    rows: List[Dict[str, object]] = []
    with p.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def merge_relation_rows(
    excel_rows: List[Dict[str, object]],
    supplement: List[Dict[str, object]] | None = None,
) -> List[Dict[str, object]]:
    supplement = supplement if supplement is not None else load_relation_supplement()
    seen = {str(r.get("relation_id") or "") for r in excel_rows if r.get("relation_id")}
    merged = list(excel_rows)
    for row in supplement:
        rid = str(row.get("relation_id") or "")
        if rid and rid in seen:
            continue
        merged.append(row)
        if rid:
            seen.add(rid)
    return merged


def chunked(rows: List[Dict[str, object]], size: int) -> Iterable[List[Dict[str, object]]]:
    for i in range(0, len(rows), size):
        yield rows[i : i + size]


def run_batches(
    session,
    query: str,
    rows: List[Dict[str, object]],
    batch_size: int,
    **params: object,
) -> int:
    total = 0
    for batch in chunked(rows, batch_size):
        session.run(query, rows=batch, **params).consume()
        total += len(batch)
    return total


def count_relationships(session) -> Dict[str, int]:
    result = session.run(
        """
        MATCH ()-[r]->()
        RETURN type(r) AS rel, count(*) AS total
        ORDER BY rel
        """
    )
    return {record["rel"]: record["total"] for record in result}


def drop_constraints_and_data(session, reset: bool) -> None:
    if not reset:
        return
    session.run("MATCH (n) DETACH DELETE n").consume()
    rows = session.run("SHOW CONSTRAINTS YIELD name RETURN name").data()
    for r in rows:
        name = str(r["name"]).replace("`", "")
        session.run(f"DROP CONSTRAINT `{name}` IF EXISTS").consume()


def create_constraints(session) -> None:
    stmts = [
        "CREATE CONSTRAINT career_code_unique IF NOT EXISTS FOR (n:Career) REQUIRE n.career_code IS UNIQUE",
        "CREATE CONSTRAINT industry_name_unique IF NOT EXISTS FOR (n:Industry) REQUIRE n.name IS UNIQUE",
        "CREATE CONSTRAINT taxonomy_name_unique IF NOT EXISTS FOR (n:Taxonomy) REQUIRE n.name IS UNIQUE",
        "CREATE CONSTRAINT comp_type_code_unique IF NOT EXISTS FOR (n:CompetencyType) REQUIRE n.type_code IS UNIQUE",
        "CREATE CONSTRAINT subject_code_unique IF NOT EXISTS FOR (n:Subject) REQUIRE n.subject_code IS UNIQUE",
        "CREATE CONSTRAINT website_code_unique IF NOT EXISTS FOR (n:Website) REQUIRE n.website_code IS UNIQUE",
        "CREATE CONSTRAINT org_code_unique IF NOT EXISTS FOR (n:Organization) REQUIRE n.org_code IS UNIQUE",
        "CREATE CONSTRAINT instructor_code_unique IF NOT EXISTS FOR (n:Instructor) REQUIRE n.instructor_code IS UNIQUE",
        "CREATE CONSTRAINT subtitle_code_unique IF NOT EXISTS FOR (n:Subtitle) REQUIRE n.subtitle_code IS UNIQUE",
        "CREATE CONSTRAINT program_code_unique IF NOT EXISTS FOR (n:Program) REQUIRE n.program_code IS UNIQUE",
        "CREATE CONSTRAINT level_code_unique IF NOT EXISTS FOR (n:Level) REQUIRE n.level_code IS UNIQUE",
        "CREATE CONSTRAINT course_code_unique IF NOT EXISTS FOR (n:Course) REQUIRE n.course_code IS UNIQUE",
        "CREATE CONSTRAINT programming_language_code IF NOT EXISTS FOR (n:ProgrammingLanguage) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT framework_code IF NOT EXISTS FOR (n:Framework) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT platform_code IF NOT EXISTS FOR (n:Platform) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT tool_code IF NOT EXISTS FOR (n:Tool) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT knowledge_code IF NOT EXISTS FOR (n:Knowledge) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT softskill_code IF NOT EXISTS FOR (n:Softskill) REQUIRE n.item_code IS UNIQUE",
        "CREATE CONSTRAINT certification_code IF NOT EXISTS FOR (n:Certification) REQUIRE n.item_code IS UNIQUE",
    ]
    for q in stmts:
        session.run(q).consume()


def merge_competency_nodes(
    session, label: str, rows: List[Dict[str, object]], batch_size: int
) -> int:
    if not rows:
        return 0
    color = NODE_COLORS.get(label, "#95A5A6")
    if label == "Tool":
        q = f"""
        UNWIND $rows AS row
        MERGE (n:{label} {{item_code: row.item_code}})
        SET n.item_name = row.item_name,
            n.color = $node_color,
            n.description = row.description
        WITH n, row
        MATCH (t:CompetencyType {{type_code: row.type_code}})
        MERGE (n)-[:OF_TYPE]->(t)
        """
    else:
        q = f"""
        UNWIND $rows AS row
        MERGE (n:{label} {{item_code: row.item_code}})
        SET n.item_name = row.item_name,
            n.color = $node_color
        WITH n, row
        MATCH (t:CompetencyType {{type_code: row.type_code}})
        MERGE (n)-[:OF_TYPE]->(t)
        """
    return run_batches(session, q, rows, batch_size, node_color=color)


def merge_careers_and_industry_taxonomy(
    session, career_rows: List[Dict[str, object]], batch_size: int
) -> int:
    q = """
    UNWIND $rows AS row
    MERGE (c:Career {career_code: row.career_code})
    SET c.career_name = row.career_name,
        c.color = $career_color
    """
    n = run_batches(
        session, q, career_rows, batch_size, career_color=NODE_COLORS["Career"]
    )

    industries = sorted(
        {
            str(r["industry"]).strip()
            for r in career_rows
            if r.get("industry") is not None and str(r["industry"]).strip()
        }
    )
    if industries:
        run_batches(
            session,
            """
            UNWIND $rows AS row
            MERGE (i:Industry {name: row.name})
            SET i.color = $color
            """,
            [{"name": x} for x in industries],
            batch_size,
            color=NODE_COLORS["Industry"],
        )

    taxonomies = sorted(
        {
            str(r["taxonomy"]).strip()
            for r in career_rows
            if r.get("taxonomy") is not None and str(r["taxonomy"]).strip()
        }
    )
    if taxonomies:
        run_batches(
            session,
            """
            UNWIND $rows AS row
            MERGE (t:Taxonomy {name: row.name})
            SET t.color = $color
            """,
            [{"name": x} for x in taxonomies],
            batch_size,
            color=NODE_COLORS["Taxonomy"],
        )

    link_ci = """
    UNWIND $rows AS row
    MATCH (c:Career {career_code: row.career_code})
    MATCH (i:Industry {name: row.industry})
    MERGE (c)-[:IN_INDUSTRY]->(i)
    """
    rows_ci = [
        {
            "career_code": r["career_code"],
            "industry": str(r["industry"]).strip(),
        }
        for r in career_rows
        if r.get("industry") is not None and str(r["industry"]).strip()
    ]
    if rows_ci:
        run_batches(session, link_ci, rows_ci, batch_size)

    link_it = """
    UNWIND $rows AS row
    MATCH (i:Industry {name: row.industry})
    MATCH (t:Taxonomy {name: row.taxonomy})
    MERGE (i)-[:IN_TAXONOMY]->(t)
    """
    rows_it = [
        {
            "industry": str(r["industry"]).strip(),
            "taxonomy": str(r["taxonomy"]).strip(),
        }
        for r in career_rows
        if r.get("industry") is not None
        and str(r["industry"]).strip()
        and r.get("taxonomy") is not None
        and str(r["taxonomy"]).strip()
    ]
    if rows_it:
        # Cặp duy nhất để tránh batch trùng
        seen = set()
        deduped = []
        for x in rows_it:
            key = (x["industry"], x["taxonomy"])
            if key not in seen:
                seen.add(key)
                deduped.append(x)
        run_batches(session, link_it, deduped, batch_size)

    return n


def merge_typed_career_competency(
    session, map_rows: List[Dict[str, object]], batch_size: int
) -> int:
    by_type: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in map_rows:
        cc = str(row.get("career_code") or "").strip()
        ic = str(row.get("item_code") or "").strip()
        if (cc, ic) in CAREER_COMPETENCY_DENYLIST:
            continue
        tc = row.get("type_code")
        if tc is None or str(tc).strip() == "":
            continue
        tc_key = str(tc).strip()
        if tc_key not in TYPE_CODE_TO_LABEL:
            continue
        by_type[tc_key].append(row)

    total = 0
    for type_code, rows in by_type.items():
        label = TYPE_CODE_TO_LABEL[type_code]
        rel = CAREER_REL_BY_TYPE[type_code]
        q = f"""
        UNWIND $rows AS row
        MATCH (career:Career {{career_code: row.career_code}})
        MATCH (comp:{label} {{item_code: row.item_code}})
        MERGE (career)-[r:{rel}]->(comp)
        SET r.priority_group = row.priority_group,
            r.map_id = row.map_id,
            r.type_code = row.type_code
        """
        total += run_batches(session, q, rows, batch_size)
    return total


def merge_competency_relations(
    session, relation_rows: List[Dict[str, object]], batch_size: int
) -> int:
    """Quan hệ năng lực–năng lực từ sheet competency_relation (BUILT_ON, VALIDATES, …)."""
    by_key: Dict[tuple[str, str, str], List[Dict[str, object]]] = defaultdict(list)
    for row in relation_rows:
        rel_type = str(row.get("relation_type") or "").strip().upper()
        from_tc = str(row.get("from_type_code") or "").strip()
        to_tc = str(row.get("to_type_code") or "").strip()
        from_ic = str(row.get("from_item_code") or "").strip()
        to_ic = str(row.get("to_item_code") or "").strip()
        if not rel_type or not from_tc or not to_tc or not from_ic or not to_ic:
            continue
        if from_tc not in TYPE_CODE_TO_LABEL or to_tc not in TYPE_CODE_TO_LABEL:
            continue
        by_key[(rel_type, from_tc, to_tc)].append(row)

    total = 0
    for (rel_type, from_tc, to_tc), rows in by_key.items():
        from_label = TYPE_CODE_TO_LABEL[from_tc]
        to_label = TYPE_CODE_TO_LABEL[to_tc]
        q = f"""
        UNWIND $rows AS row
        MATCH (src:{from_label} {{item_code: row.from_item_code}})
        MATCH (dst:{to_label} {{item_code: row.to_item_code}})
        MERGE (src)-[r:{rel_type}]->(dst)
        SET r.relation_id = row.relation_id,
            r.from_type_code = row.from_type_code,
            r.to_type_code = row.to_type_code,
            r.note = row.note
        """
        total += run_batches(session, q, rows, batch_size)
    return total


def merge_typed_course_competency(
    session, map_rows: List[Dict[str, object]], batch_size: int
) -> int:
    by_type: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in map_rows:
        rt = str(row.get("relation_type") or "TEACH").strip().upper()
        if rt != "TEACH":
            continue
        tc = row.get("type_code")
        if tc is None or str(tc).strip() == "":
            continue
        tc_key = str(tc).strip()
        if tc_key not in TYPE_CODE_TO_LABEL:
            continue
        by_type[tc_key].append(row)

    total = 0
    for type_code, rows in by_type.items():
        label = TYPE_CODE_TO_LABEL[type_code]
        rel = TEACH_REL_BY_TYPE[type_code]
        q = f"""
        UNWIND $rows AS row
        MATCH (course:Course {{course_code: row.course_code}})
        MATCH (comp:{label} {{item_code: row.item_code}})
        MERGE (course)-[r:{rel}]->(comp)
        SET r.relation_type = row.relation_type,
            r.coverage_level = row.coverage_level,
            r.map_id = row.map_id,
            r.type_code = row.type_code
        """
        total += run_batches(session, q, rows, batch_size)
    return total


def ingest_xlsx_to_neo4j(
    xlsx_path: Path, batch_size: int, *, reset_graph: bool = False
) -> None:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    sheet_cache: Dict[str, List[Dict[str, object]]] = {
        "career": read_sheet_rows(xlsx_path, "career"),
        "competency_type": read_sheet_rows(xlsx_path, "competency_type"),
        "career_competency_map": read_sheet_rows(xlsx_path, "career_competency_map"),
        "competency_relation": merge_relation_rows(
            read_sheet_rows(xlsx_path, "competency_relation")
        ),
        "subject": read_sheet_rows(xlsx_path, "subject"),
        "website": read_sheet_rows(xlsx_path, "website"),
        "organization": read_sheet_rows(xlsx_path, "organization"),
        "instructor": read_sheet_rows(xlsx_path, "instructor"),
        "subtitle": read_sheet_rows(xlsx_path, "subtitle"),
        "program": read_sheet_rows(xlsx_path, "program"),
        "level": read_sheet_rows(xlsx_path, "level"),
        "course": read_sheet_rows(xlsx_path, "course"),
        "course_competency_map": read_sheet_rows(xlsx_path, "course_competency_map"),
    }
    for sheet_name, _ in COMPETENCY_SHEETS:
        sheet_cache[sheet_name] = read_sheet_rows(xlsx_path, sheet_name)

    uri = os.getenv("NEO4J_URI", "bolt://localhost:7687")
    user = os.getenv("NEO4J_USER", "neo4j")
    password = os.getenv("NEO4J_PASSWORD", "neo4j_password")

    driver = GraphDatabase.driver(uri, auth=(user, password))

    with driver.session() as session:
        drop_constraints_and_data(session, reset_graph)
        create_constraints(session)

        totals: Dict[str, int] = {}

        ct_color = NODE_COLORS["CompetencyType"]
        totals["CompetencyType"] = run_batches(
            session,
            """
            UNWIND $rows AS row
            MERGE (t:CompetencyType {type_code: row.type_code})
            SET t.type_name = row.type_name,
                t.color = $tc_color
            """,
            sheet_cache["competency_type"],
            batch_size,
            tc_color=ct_color,
        )

        for sheet_name, label in COMPETENCY_SHEETS:
            key = f"Skill:{label}"
            totals[key] = merge_competency_nodes(
                session, label, sheet_cache[sheet_name], batch_size
            )

        totals["CompetencyRelation"] = merge_competency_relations(
            session, sheet_cache["competency_relation"], batch_size
        )

        totals["Career"] = merge_careers_and_industry_taxonomy(
            session, sheet_cache["career"], batch_size
        )

        totals["CareerCompetency"] = merge_typed_career_competency(
            session, sheet_cache["career_competency_map"], batch_size
        )

        def _color_batch(
            key: str, cypher: str, rows: List[Dict[str, object]], color_key: str
        ) -> None:
            totals[key] = run_batches(
                session, cypher, rows, batch_size, node_color=NODE_COLORS[color_key]
            )

        _color_batch(
            "Subject",
            """
            UNWIND $rows AS row
            MERGE (s:Subject {subject_code: row.subject_code})
            SET s.subject_name = row.subject_name, s.color = $node_color
            """,
            sheet_cache["subject"],
            "Subject",
        )
        _color_batch(
            "Website",
            """
            UNWIND $rows AS row
            MERGE (w:Website {website_code: row.website_code})
            SET w.website_name = row.website_name,
                w.website_url = row.website_url,
                w.color = $node_color
            """,
            sheet_cache["website"],
            "Website",
        )
        _color_batch(
            "Organization",
            """
            UNWIND $rows AS row
            MERGE (o:Organization {org_code: row.org_code})
            SET o.org_name = row.org_name, o.color = $node_color
            """,
            sheet_cache["organization"],
            "Organization",
        )
        _color_batch(
            "Instructor",
            """
            UNWIND $rows AS row
            MERGE (i:Instructor {instructor_code: row.instructor_code})
            SET i.instructor_name = row.instructor_name, i.color = $node_color
            WITH i, row
            OPTIONAL MATCH (o:Organization {org_code: row.org_code})
            FOREACH (_ IN CASE WHEN o IS NULL THEN [] ELSE [1] END | MERGE (i)-[:BELONGS_TO]->(o))
            """,
            sheet_cache["instructor"],
            "Instructor",
        )
        _color_batch(
            "Subtitle",
            """
            UNWIND $rows AS row
            MERGE (s:Subtitle {subtitle_code: row.subtitle_code})
            SET s.subtitle_name = row.subtitle_name, s.color = $node_color
            """,
            sheet_cache["subtitle"],
            "Subtitle",
        )
        _color_batch(
            "Program",
            """
            UNWIND $rows AS row
            MERGE (p:Program {program_code: row.program_code})
            SET p.program_name = row.program_name, p.color = $node_color
            """,
            sheet_cache["program"],
            "Program",
        )
        _color_batch(
            "Level",
            """
            UNWIND $rows AS row
            MERGE (l:Level {level_code: row.level_code})
            SET l.level_name = row.level_name, l.color = $node_color
            """,
            sheet_cache["level"],
            "Level",
        )

        course_q = """
        UNWIND $rows AS row
        MERGE (c:Course {course_code: row.course_code})
        SET c.course_name = row.course_name,
            c.duration_hours = row.duration_hours,
            c.course_type = row.course_type,
            c.language = row.language,
            c.url = row.url,
            c.description = row.description,
            c.color = $node_color
        WITH c, row
        OPTIONAL MATCH (s:Subject {subject_code: row.subject_code})
        FOREACH (_ IN CASE WHEN s IS NULL THEN [] ELSE [1] END | MERGE (c)-[:IN_SUBJECT]->(s))
        WITH c, row
        OPTIONAL MATCH (p:Program {program_code: row.program_code})
        FOREACH (_ IN CASE WHEN p IS NULL THEN [] ELSE [1] END | MERGE (c)-[:IN_PROGRAM]->(p))
        WITH c, row
        OPTIONAL MATCH (l:Level {level_code: row.level_code})
        FOREACH (_ IN CASE WHEN l IS NULL THEN [] ELSE [1] END | MERGE (c)-[:AT_LEVEL]->(l))
        WITH c, row
        OPTIONAL MATCH (w:Website {website_code: row.website_code})
        FOREACH (_ IN CASE WHEN w IS NULL THEN [] ELSE [1] END | MERGE (c)-[:ON_WEBSITE]->(w))
        WITH c, row
        OPTIONAL MATCH (o:Organization {org_code: row.org_code})
        FOREACH (_ IN CASE WHEN o IS NULL THEN [] ELSE [1] END | MERGE (c)-[:PROVIDED_BY]->(o))
        WITH c, row
        OPTIONAL MATCH (i:Instructor {instructor_code: row.instructor_code})
        FOREACH (_ IN CASE WHEN i IS NULL THEN [] ELSE [1] END | MERGE (c)-[:TAUGHT_BY]->(i))
        WITH c, row
        OPTIONAL MATCH (sub:Subtitle {subtitle_code: row.subtitle_code})
        FOREACH (_ IN CASE WHEN sub IS NULL THEN [] ELSE [1] END | MERGE (c)-[:HAS_SUBTITLE]->(sub))
        """
        totals["Course"] = run_batches(
            session,
            course_q,
            sheet_cache["course"],
            batch_size,
            node_color=NODE_COLORS["Course"],
        )

        totals["CourseCompetency"] = merge_typed_course_competency(
            session, sheet_cache["course_competency_map"], batch_size
        )

        rel_counts = count_relationships(session)

    driver.close()

    print("Ingest xong. Số dòng / thao tác:")
    for key, count in totals.items():
        print(f"  - {key}: {count}")
    print("Số quan hệ theo type:")
    for rel, total in rel_counts.items():
        print(f"  * {rel}: {total}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Nạp Excel vào Neo4j (mô hình NEED_*/TEACH_*).")
    p.add_argument("--xlsx-path", type=Path, default=DEFAULT_XLSX_PATH)
    p.add_argument("--batch-size", type=int, default=500)
    p.add_argument(
        "--reset-graph",
        action="store_true",
        help="Xóa toàn bộ nút/cạnh và drop constraints trước khi nạp (sạch DB).",
    )
    return p.parse_args()


def main() -> None:
    load_dotenv(PROJECT_ROOT / ".env", override=True)
    args = parse_args()
    ingest_xlsx_to_neo4j(args.xlsx_path, args.batch_size, reset_graph=args.reset_graph)


if __name__ == "__main__":
    main()
