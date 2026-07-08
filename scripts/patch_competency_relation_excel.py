"""
Bổ sung cạnh REQUIRES (softskill) và REQUIRES_KNOWLEDGE (methodology) vào sheet competency_relation.

Chạy: python scripts/patch_competency_relation_excel.py
Sau đó: python scripts/ingest.py --reset-graph
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
XLSX = PROJECT_ROOT / "data" / "bộ dữ liệu.xlsx"
SHEET = "competency_relation"

NEW_ROWS: list[dict[str, str]] = [
    # REQUIRES softskill
    {"relation_id": "CR027", "from_item_code": "S_LEAD", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Leadership cần Communication"},
    {"relation_id": "CR028", "from_item_code": "S_COACH", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Coaching cần Communication"},
    {"relation_id": "CR029", "from_item_code": "S_COACH", "from_type_code": "CT_SOFT", "to_item_code": "S_LEAD", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Coaching cần Leadership"},
    {"relation_id": "CR030", "from_item_code": "S_NEGO", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Negotiation cần Communication"},
    {"relation_id": "CR031", "from_item_code": "S_FACI", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Facilitation cần Communication"},
    {"relation_id": "CR032", "from_item_code": "S_STAKE", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Stakeholder Management cần Communication"},
    {"relation_id": "CR033", "from_item_code": "S_ARCH_DECISION", "from_type_code": "CT_SOFT", "to_item_code": "S_CRIT", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Architectural Decision cần Critical Thinking"},
    {"relation_id": "CR034", "from_item_code": "S_RISK_MGMT", "from_type_code": "CT_SOFT", "to_item_code": "S_PROB", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Risk Management cần Problem Solving"},
    {"relation_id": "CR035", "from_item_code": "S_LEAD", "from_type_code": "CT_SOFT", "to_item_code": "S_COLLAB", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Leadership cần Collaboration"},
    {"relation_id": "CR036", "from_item_code": "S_DOC", "from_type_code": "CT_SOFT", "to_item_code": "S_COMM", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Documentation cần Communication"},
    {"relation_id": "CR037", "from_item_code": "S_TIME", "from_type_code": "CT_SOFT", "to_item_code": "S_ADAPT", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Time Management cần Adaptability"},
    {"relation_id": "CR038", "from_item_code": "S_STAKE", "from_type_code": "CT_SOFT", "to_item_code": "S_NEGO", "to_type_code": "CT_SOFT", "relation_type": "REQUIRES", "note": "Stakeholder Management cần Negotiation"},
    # REQUIRES_KNOWLEDGE methodology
    {"relation_id": "CR039", "from_item_code": "F_AGILE", "from_type_code": "CT_FRAM", "to_item_code": "K_SDLC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "Agile cần hiểu SDLC"},
    {"relation_id": "CR040", "from_item_code": "F_KANBAN", "from_type_code": "CT_FRAM", "to_item_code": "K_PROC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "Kanban cần Process Analysis"},
    {"relation_id": "CR041", "from_item_code": "F_BABOK", "from_type_code": "CT_FRAM", "to_item_code": "K_REQ", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "BABOK cần Requirements Engineering"},
    {"relation_id": "CR042", "from_item_code": "F_BPMN", "from_type_code": "CT_FRAM", "to_item_code": "K_PROC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "BPMN cần Process Analysis"},
    {"relation_id": "CR043", "from_item_code": "F_UML", "from_type_code": "CT_FRAM", "to_item_code": "K_OOP", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "UML cần OOP"},
    {"relation_id": "CR044", "from_item_code": "F_RACI", "from_type_code": "CT_FRAM", "to_item_code": "K_PROC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "RACI cần Process Analysis"},
    {"relation_id": "CR045", "from_item_code": "F_OKR", "from_type_code": "CT_FRAM", "to_item_code": "K_PROC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "OKR/KPI cần Process Analysis"},
    {"relation_id": "CR046", "from_item_code": "F_DTHINK", "from_type_code": "CT_FRAM", "to_item_code": "K_OOP", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "Design Thinking cần OOP and Design Principles"},
    {"relation_id": "CR047", "from_item_code": "F_SAFE", "from_type_code": "CT_FRAM", "to_item_code": "K_SDLC", "to_type_code": "CT_KNOW", "relation_type": "REQUIRES_KNOWLEDGE", "note": "SAFe cần SDLC"},
]


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Missing {XLSX}")

    wb = load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET} not found")

    ws = wb[SHEET]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    existing_ids: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        item = dict(zip(headers, row))
        rid = str(item.get("relation_id") or "").strip()
        if rid:
            existing_ids.add(rid)

    added = 0
    for row in NEW_ROWS:
        if row["relation_id"] in existing_ids:
            continue
        ws.append([row.get(h, "") for h in headers])
        added += 1

    wb.save(XLSX)
    print(f"OK: added {added} rows to {SHEET} (skipped existing relation_id)")


if __name__ == "__main__":
    main()
